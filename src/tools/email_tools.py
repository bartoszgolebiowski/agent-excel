from __future__ import annotations

import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Font

from .models import (
    ArchiveEmailRequest,
    ArchiveEmailResponse,
    CheckInboxRequest,
    CheckInboxResponse,
    ReadEmailRequest,
    ReadEmailResponse,
    SaveToExcelRequest,
    SaveToExcelResponse,
)


def check_inbox(request: CheckInboxRequest) -> CheckInboxResponse:
    """Check inbox folder for email files."""
    inbox_path = Path(request.inbox_path)

    if not inbox_path.exists():
        inbox_path.mkdir(parents=True, exist_ok=True)
        return CheckInboxResponse(files=[], count=0)

    # Get all .txt files in the inbox
    files = [str(f) for f in inbox_path.glob("*.txt")]
    files.sort()  # Process in consistent order

    return CheckInboxResponse(files=files, count=len(files))


def read_email(request: ReadEmailRequest) -> ReadEmailResponse:
    """Read the content of an email file."""
    file_path = Path(request.file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"Email file not found: {request.file_path}")

    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    return ReadEmailResponse(content=content, file_name=file_path.name)


def save_to_excel(request: SaveToExcelRequest) -> SaveToExcelResponse:
    """Append email analysis to Excel report."""
    excel_path = Path(request.excel_path)

    # Create parent directory if it doesn't exist
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    # Load or create workbook
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        assert ws is not None, "Workbook must have an active sheet"
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None, "New workbook must have an active sheet"
        ws.title = "Email Analysis"

        # Create header row
        headers = [
            "Main Topic",
            "Business Category",
            "Contact Data",
            "Urgency",
            "Sentiment",
            "Summary",
            "Event Date",
            "Source File",
        ]
        ws.append(headers)

        # Style header row
        for cell in ws[1]:
            cell.font = Font(bold=True)

    # Append new row
    row_data = [
        request.main_topic,
        request.business_category,
        request.contact_data,
        request.urgency,
        request.sentiment,
        request.summary,
        request.event_date,
        request.source_file,
    ]
    ws.append(row_data)

    # Save workbook
    wb.save(excel_path)

    return SaveToExcelResponse(
        success=True, message=f"Successfully saved analysis to {excel_path}"
    )


def archive_email(request: ArchiveEmailRequest) -> ArchiveEmailResponse:
    """Move processed email to archive folder."""
    source_path = Path(request.file_path)
    archive_path = Path(request.archive_path)

    if not source_path.exists():
        raise FileNotFoundError(f"Source file not found: {request.file_path}")

    # Create archive directory if it doesn't exist
    archive_path.mkdir(parents=True, exist_ok=True)

    # Move file to archive
    destination = archive_path / source_path.name

    # Handle duplicate names by adding timestamp
    if destination.exists():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = source_path.stem
        suffix = source_path.suffix
        destination = archive_path / f"{stem}_{timestamp}{suffix}"

    shutil.move(str(source_path), str(destination))

    return ArchiveEmailResponse(
        success=True,
        message=f"Successfully archived {source_path.name} to {destination}",
    )
